#!/usr/bin/env python
# coding: utf-8

# In[1]:


import datetime
import traceback
import os
import requests
import time
import json
import pyodbc
import textwrap
import pytz
import pandas as pd
import math
from openpyxl import load_workbook
import os
import datetime
from pathlib import Path


# In[2]:


# driver
driver = '{ODBC Driver 17 for SQL Server}'


# In[3]:


# server and db name
server_name = 'dublinbus-team12-server'
database_name = 'dublinbus-team12-db'


# In[4]:


# server
server = '{}.database.windows.net,1433'.format(server_name)


# In[5]:


# username and passord 
username = 'innovationgeeks'
password = 'Laurawillsaveus445!'


# In[6]:


# connection string 
connection_string = textwrap.dedent('''
    Driver={driver};
    Server={server};
    Database={database};
    Uid={username};
    Pwd={password};
    Encrypt=yes;
    TrustServerCertificate=no;
    Connection Timeout=30;
'''.format(
    driver = driver, 
    server = server, 
    database = database_name,
    username = username,
    password = password
    )                               
)


# In[7]:


# pyodbc conncetion object
# conn: pyodbc.Connection = pyodbc.connect(connection_string)


# In[ ]:


# curser object for connection 
# cursor: pyodbc.Cursor = conn.cursor()


# In[6]:


server = "dublinbus-team12-server.database.windows.net"
db = "dublinbus-team12-db"
user = "innovationgeeks"
password = "Laurawillsaveus445!"
port = "1433"
conn = pyodbc.connect('DRIVER={SQL Server};SERVER=' + server + ';PORT=' + port + ';DATABASE=' + db +';UID=' + user + ';PWD=' + password)
cursor = conn.cursor()


# In[10]:


gtfsAndStops = '''
    CREATE TABLE GTFSWithTraffic(
        now VARCHAR(256), 
        bus_Number VARCHAR(256),
        I_O VARCHAR(256), 
        stop_sequence_no INT,
        stop_id VARCHAR(256), 
        stop_lat FLOAT, 
        stop_lng FLOAT,
        Arr_Delay FLOAT,
        curr_speed INT,
        freeflow_speed INT,
        curr_travel_time INT,
        freeflow_travel_time INT,
        pressure FLOAT,
        cloud INT,
        rain INT,
        humidity FLOAT
    )
'''

# stop_name TEXT,


# In[11]:


def create_table():
    '''function to create table in azure db - only needs to be created once'''
    try:
        cursor.execute(gtfsAndStops)
        conn.commit()
    except Exception as e:
        print(e)


# In[12]:


# # test create
create_table()


# In[13]:


# Drop table if needed
# cursor.execute("DROP TABLE dbo.GTFSWithTraffic")
# conn.commit()


# In[29]:


# Bus Routs 
br = "1,4,7,9,11,13,14,15,16,17,18,25,26,27,31,32,33,37,38,39,40,41,42,43,44,47,49,53,59,61,63,65,66,67,68,69,70,75,76,79,83,84,102,104,111,114,116,118,120,122,123,130,130,140,142,145,150,151,161,184,185,220,236,238,239,270,14C,15A,15B,15D,16C,16D,17A,25A,25B,25D,25X,27A,27B,27X,29A,31A,31B,31D,32X,33A,33B,33D,33E,33X,38A,38B,38D,39A,39X,40B,40D,40E,41A,41B,41C,41D,41X,42D,44B,45A,46A,46E,51D,51X,54A,56A,65B,66A,66B,66X,67X,68A,68X,69X,70D,76A,77A,77X,79A,7A,7B,7D,83A,84A,84X"
BusRouts = br.split(",")


# In[24]:


# url = 'https://api.nationaltransport.ie/gtfsrtest/?format=json'
# headers = {'x-api-key' : '35665880db8e44d381ff2efde4d92628'}
# r = requests.get(url, headers=headers)
# data = json.loads(r.text)
# print(data)


# In[30]:


def get_gtfs_data(): # data in parameters
    ''' 
    Gets data from gtfs data
    '''
    # now variable 
    tz = pytz.timezone('Europe/Dublin')
    time = datetime.datetime.now(tz=tz)
    now = time.strftime('%Y-%m-%d %H:%M:%S')
    
    url = 'https://api.nationaltransport.ie/gtfsrtest/?format=json'
    headers = {'x-api-key' : '082521757ec741f0b414ae368c1720ff'}
    r = requests.get(url, headers=headers)
    data = json.loads(r.text)
    
    # loop through data to find just dublin bus routs
    Output = []
    for bus in data['entity']:
        routeSpecific = []
        busNumber = bus.get("id").split("-")[1]
        
        # if bus number is in busRouts
        if busNumber in BusRouts:
            routeSpecific.append(now) 
            routeSpecific.append(bus.get("id").split("-")[1])
            routeSpecific.append(bus.get("id").split("-")[3][-1:])
            routeSpecific.append(bus.get("trip_update").get("trip").get("start_time"))
            routeSpecific.append(bus.get("trip_update").get("trip").get("start_date"))
            routeSpecific.append(bus.get("trip_update").get("trip").get("route_id"))
            
            # loop through list of stops to get stop id, delay for each bustop 
            try:
                for i in data['entity'][data['entity'].index(bus)]['trip_update']['stop_time_update']:
                    stopSpecific = []
                    stopSpecific.append(i["stop_sequence"])
                    stopSpecific.append(i["stop_id"])
                    try:
                        stopSpecific.append(i["arrival"]["delay"])    
                    except:
                        stopSpecific.append(0)
                    try:
                        stopSpecific.append(i["departure"]["delay"]) 
                    except:
                        stopSpecific.append(0)
                    finalProduct = routeSpecific+stopSpecific
                    Output.append(finalProduct)
            except:
                pass
    return Output


# In[31]:


def gtfs_to_stops(func_data):
    '''
    takes gtfs data and links stops data - returns tuple of values 
    '''
    df_gtfs = pd.DataFrame(func_data, columns =['DateTime', 'Bus_No', 'Inbound', 'Time_Dep', 'Date', 'Trip_ID', 'Stop_No','Stop_ID','Arr_Delay','Dep_Delay'])
    
    stops_data = pd.read_excel(r"stops_data.xlsx")
    df_stops = pd.DataFrame(stops_data, columns= ['stop_id', 'stop_name', 'stop_lat', 'stop_lon'])
    
    # df_gtfs['stop_name'] = 'nan'
    df_gtfs['stop_lat'] = 0.0
    df_gtfs['stop_lon'] = 0.0
    

    for i in df_gtfs.index:
        for j in df_stops.index:
            if(df_gtfs['Stop_ID'][i] == df_stops['stop_id'][j]):
                # df_gtfs['stop_name'][i] = df_stops['stop_name'][j]
                df_gtfs['stop_lat'][i] = df_stops['stop_lat'][j]
                df_gtfs['stop_lon'][i] = df_stops['stop_lon'][j]
                
    tuple_df = list(df_gtfs.itertuples(index=False, name = None))
    
    return tuple_df


# In[32]:


def Excel_append():
    func_data = get_gtfs_data()
    stops = gtfs_to_stops(func_data)
    df_temp = pd.DataFrame(stops, columns =['now', 'busNumber', 'I_O', 'start_time', 'start_date', 'route_id', 'stop_sequence_number', 'stop_id', 'Arr_Delay', 'Dep_Delay', 'stop_lat', 'stop_lon'])

    try:
        book = load_workbook('Data_30_min_GTFS.xlsx')
        writer = pd.ExcelWriter('Data_30_min_GTFS.xlsx', engine='openpyxl') 
        writer.book = book

        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        # df_rand_stops.to_excel(writer, "Sheet1")
        df_temp.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, index = False,header= False)
        writer.save()
    except:
        writer = pd.ExcelWriter('Data_30_min_GTFS.xlsx')
        df_temp.to_excel(writer, index=False)
        writer.save()


# In[33]:


#Excel_append()


# In[ ]:


#df_rand_stops = pd.read_excel(r'Bus_Selected_Routes.xlsx', index_col=0)


# In[ ]:


#df_rand_stops['Arr_Delay'] = 0


# In[ ]:


#df_GTFS = pd.read_excel (r'Data_30_min_GTFS.xlsx')


# In[ ]:


#df_merge = pd.merge(df_GTFS, df_rand_stops, how='left', left_on= ["busNumber",'I_O','stop_id'], right_on = ["Bus_No",'IO','Stop_ID'])


# In[ ]:


#df_selected = df_merge[df_merge.isnull().any(axis=1) == False]


# In[ ]:


#df_selected


# In[53]:


def Data_With_Delay_Merge():
    df_rand_stops = pd.read_excel(r'Bus_Selected_Routes_1_stop.xlsx', index_col=0)
    df_rand_stops['Arr_Delay'] = 0
    df_GTFS = pd.read_excel (r'Data_30_min_GTFS.xlsx')
    df_merge = pd.merge(df_GTFS, df_rand_stops, how='left', left_on= ["busNumber",'I_O','stop_id'], right_on = ["Bus_No",'IO','Stop_ID'])
    df_selected = df_merge[df_merge.isnull().any(axis=1) == False]
    try:
        for index, stop in df_rand_stops.iterrows():
            for i, r in df_selected.iterrows():
                if((stop.Stop_ID == r.stop_id)):
                    df_rand_stops['Arr_Delay'][index] = r.Arr_Delay_x
        my_file = Path("Temp_dataframe.xlsx")
        if my_file.is_file():
            os.remove('Temp_dataframe.xlsx')
        writer = pd.ExcelWriter('Temp_dataframe.xlsx')
        df_rand_stops.to_excel(writer)
        writer.save()
    except Exception as e:
        print("this is the row", row)
        print(e)


# In[54]:


#Data_With_Delay_Merge()


# In[55]:


def tomtom(lat, lon):
    r = requests.get("https://api.tomtom.com/traffic/services/4/flowSegmentData/relative0/10/json?point={}%2C{}&unit=KMPH&key=Oifx1JdvaGcExvXiqPmtVNfQIAR3j0of".format(lat, lon))
    data = json.loads(r.text)
    vals = data.get("flowSegmentData").get("currentSpeed"), data.get("flowSegmentData").get("freeFlowSpeed"), data.get("flowSegmentData").get("currentTravelTime"),data.get("flowSegmentData").get("freeFlowTravelTime")
    
    return vals


# In[56]:


def binder():
    df_rand_stops = pd.read_excel(r'Temp_dataframe.xlsx', index_col=0)
    df_rand_stops['CurrentSpeed'] = 0
    df_rand_stops['FreeFlowSpeed'] = 0
    df_rand_stops['CurrentTravelTime'] = 0
    df_rand_stops['FreeFlowTravelTime'] = 0
    
    count = 0
    for index, delay in df_rand_stops.iterrows():
        if(delay.Stop_Lat != 0 and delay.Stop_Lng != 0):
            values = tomtom(delay.Stop_Lat, delay.Stop_Lng)
            df_rand_stops['CurrentSpeed'][index] = values[0]
            df_rand_stops['FreeFlowSpeed'][index] = values[1]
            df_rand_stops['CurrentTravelTime'][index] = values[2]
            df_rand_stops['FreeFlowTravelTime'][index] = values[3]
            count = count + 1

    my_file = Path("Temp_dataframe.xlsx")
    if my_file.is_file():
        os.remove('Temp_dataframe.xlsx')
    writer = pd.ExcelWriter('Temp_dataframe.xlsx')
    df_rand_stops.to_excel(writer)
    writer.save()


# In[57]:





# In[58]:


def Weather():
    df_rand_stops = pd.read_excel(r'Temp_dataframe.xlsx', index_col=0)
    cursor = conn.cursor()
    Weather_Query = """select TOP(1) * from [dbo].[weather_current] ORDER By [now] DESC"""
    cursor.execute(Weather_Query)
    data = cursor.fetchall()
    df_rand_stops['Pressure'] = data[0][7]
    df_rand_stops['Cloud'] = data[0][3]
    df_rand_stops['Rain'] = data[0][4]
    df_rand_stops['Humidity'] = data[0][8]
    
    tz = pytz.timezone('Europe/Dublin')
    time = datetime.datetime.now(tz=tz)
    now = time.strftime('%Y-%m-%d %H:%M:%S')
    df_rand_stops['now'] = now
    first_column = df_rand_stops.pop('now')
    df_rand_stops.insert(0, 'now', first_column)
    All_Data = list(df_rand_stops.itertuples(index=False, name=None))
    os.remove('Temp_dataframe.xlsx')

    return All_Data


# In[59]:


def push_to_db():
#     func_data = get_gtfs_data()
#     stops = gtfs_to_stops(func_data)
#     print("this is the len of stops", len(stops))
    stops = Weather()
    try: 
        for row in stops:
            cursor.execute("""INSERT INTO dbo.GTFSWithTraffic(now, bus_Number, I_O, stop_sequence_no, stop_id, stop_lat, stop_lng, Arr_Delay, curr_speed, freeflow_speed, curr_travel_time, freeflow_travel_time, pressure, cloud, rain, humidity) VALUES{}""".format(row))
        conn.commit()
    except Exception as e:
        print("this is the row", row)
        print(e)


# In[60]:




# In[31]:


# to check if the data is getting as expected - returns xl sheet with output
#writer = pd.ExcelWriter('Final_test.xlsx')
# df_rand_stops.to_excel(writer)
# writer.save()


# In[ ]:


# # delete the excel file after 1 hr 
# os.remove('Data_30_min_GTFS.xlsx')


# In[ ]:


def main():        
    try:
        Excel_append()
        Data_With_Delay_Merge()
        binder()
        push_to_db()
    except Exception as e:
        print(e)
        time.sleep(60)
        print(traceback.format_exc)
        print("error found while querying")

if __name__ == "__main__":
    main()


# to do
# - get return values in tuple 
# - add other model parameters
# - make push to db function 
# - main function working 
# - duplicate 
# - cron 





# %%
